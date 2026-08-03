#!/usr/bin/env python3
"""Validate publication invariants before an automated checkpoint commit."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]


def load_json(path: Path) -> dict:
    with path.open(encoding="utf-8") as handle:
        payload = json.load(handle)
    if not isinstance(payload, dict):
        raise ValueError(f"{path} must contain a JSON object")
    return payload


def main() -> int:
    data_news = load_json(ROOT / "data" / "news.json")
    public_news = load_json(ROOT / "docs" / "data" / "news.json")
    if data_news != public_news:
        raise ValueError("data/news.json and docs/data/news.json differ")
    site_news = load_json(ROOT / "docs" / "data" / "news-lite.json")

    items = public_news.get("items", [])
    if public_news.get("article_count") != len(items):
        raise ValueError("article_count does not match the JSON item count")
    item_ids = [str(item.get("id") or "") for item in items]
    if not all(item_ids) or len(item_ids) != len(set(item_ids)):
        raise ValueError("public article IDs are empty or duplicated")
    if any(item.get("source") == "OpenAI News" for item in items):
        raise ValueError("retired OpenAI News records remain public")
    chunk_names = site_news.get("chunks", [])
    if not isinstance(chunk_names, list) or not all(
        isinstance(name, str) and name.startswith("news-lite-")
        for name in chunk_names
    ):
        raise ValueError("site JSON chunk manifest is invalid")
    site_items = [
        item
        for chunk_name in chunk_names
        for item in load_json(ROOT / "docs" / "data" / chunk_name).get("items", [])
    ]
    site_item_ids = [str(item.get("id") or "") for item in site_items]
    if site_news.get("article_count") != len(site_items):
        raise ValueError("site article_count does not match the JSON item count")
    if site_item_ids != item_ids:
        raise ValueError("site JSON article IDs/order do not match public JSON")
    if any("summary_original" in item for item in site_items):
        raise ValueError("site JSON contains oversized original summaries")

    config = load_json(ROOT / "config" / "sources.json")
    active_sources = [
        source for source in config.get("sources", []) if source.get("active")
    ]
    source_names = [str(source.get("name") or "") for source in active_sources]
    expected_source_count = int(
        config.get("expected_active_source_count", len(source_names))
    )
    if len(source_names) != expected_source_count:
        raise ValueError(
            f"expected {expected_source_count} active sources, "
            f"found {len(source_names)}"
        )
    if not all(source_names) or len(source_names) != len(set(source_names)):
        raise ValueError("active source names are empty or duplicated")
    if "OpenAI News" in source_names:
        raise ValueError("OpenAI News remains active")
    if public_news.get("source_count") != expected_source_count:
        raise ValueError(
            "source_count does not match expected_active_source_count"
        )

    workbook = load_workbook(
        ROOT / "docs" / "innovation_news_ledger.xlsx",
        read_only=True,
        data_only=False,
    )
    ledger = workbook["News Ledger"]
    ledger_ids = [
        str(row[0] or "")
        for row in ledger.iter_rows(
            min_row=4,
            min_col=12,
            max_col=12,
            values_only=True,
        )
        if row[0]
    ]
    if ledger_ids != item_ids:
        raise ValueError("Excel article IDs/order do not match public JSON")

    registry = workbook["Source Registry"]
    registry_names = [
        str(row[0] or "")
        for row in registry.iter_rows(
            min_row=4,
            min_col=5,
            max_col=5,
            values_only=True,
        )
        if row[0]
    ]
    if registry_names != source_names:
        raise ValueError("Excel source registry does not match config order")

    print(
        json.dumps(
            {
                "status": "ok",
                "articles": len(item_ids),
                "sources": len(source_names),
                "openai_news": 0,
                "json_excel_ids_match": True,
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
